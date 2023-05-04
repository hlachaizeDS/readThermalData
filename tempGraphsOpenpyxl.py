from exportCSV import *
from tkinter import *
from tkinter import filedialog
from time import *
from smallFunctions import *
import openpyxl
from os import path

#from xlrd import open_workbook
import xlwt
#from xlutils.copy import copy
import os.path

def getFolderPath():
    root = Tk()
    root.withdraw()
    path = filedialog.askopenfilename(initialdir="C:\\Users\\HL\\Desktop\\Prototype\\Thermal_Camera")
    filename=path.split("/")[-1]
    return path[:-len(filename)-3]


if __name__ == "__main__":

    initial_path=getFolderPath()
    print(initial_path)

    typeOfPictures=".png"

    steps=['AftPremixInc']
    #steps=['AFTER_ELONGATION_DISPENSE','AFTER_ELONGATION_HEATING_SHAKING']

    #wells=[["Left up",11,46],["Left middle",42,45],["Left down",83,46],["Left-1 up",11,55],["Left-1 middle",41,55],["Left-1 down",83,56],["Middle up",10,66],["Middle middle",44,93],["Middle down",85,95],["Right up",12,123],["Right middle",43,124],["Right down",85,125]]
    wells=[]
    A1=[int(107/3),int(51/3)]
    H1=[int(104/3),int(165/3)]
    A12=[int(288/3),int(52/3)]
    #A1=[42,20]
    #H1=[41,78]
    #A12=[133,22]
    firstWell=1
    lastWell=96
    for well in range(firstWell,lastWell+1):
        coordinates=returnXY(well,A1,H1,A12)
        wells.append([getWellName(well),coordinates[1],coordinates[0]])

    #Clogging test : easier to just give all wells
    wells=[["B3",106,117],["D3",107,193],["G3",110,298],["C5",180,154],["E5",180,229],["B7",254,117],["D7",254,193],["G7",254,302],["C9",329,155],["E9",329,230]]
    for well in wells:
        buff=well[1]
        well[1]=int(well[2]/3)
        well[2] = int(buff / 3)

    temp_excel = openpyxl.Workbook()
    sheets=[]
    for step in steps:
        sheets.append(temp_excel.create_sheet(title=step))
        sheets[steps.index(step)]["A1"]= "Cycle"
        for well in wells:
            sheets[steps.index(step)].cell(row=1,column=2+wells.index(well)).value=well[0]+"("+str(well[1])+";"+str(well[2])+")"

    subdirs = [(x[0].replace("/","\\")).split("\\")[-1] for x in os.walk(initial_path)]
    subdirs.pop(0) #we remove first element as it's the initial folder
    nb_cycles=max(int(x) for x in subdirs)
    cycle_min=min(int(x) for x in subdirs)

    for cycle in range(1,nb_cycles+1):
        folderPath=initial_path+ "\\" +str(cycle)
        files=[x for x in os.listdir(folderPath)]
        print(files)
        for step in steps:
            for file in files:
                if (file.find(step)>-1) & (file.find(typeOfPictures)>-1) :
                    pathString=folderPath+"\\"+file
                    outputString = folderPath + "\\" + file[:len(file) - len(typeOfPictures)] + ".csv"
                    if path.exists(outputString)!=1:
                        openFile(pathString.replace("/","\\"))
                        exportToCsv(outputString.replace("/","\\"))
                    sleep(0.5)
                    tempMatrix=readCsv(outputString.replace("/","\\"))

                    sheets[steps.index(step)].cell(row=1+cycle,column=1).value=cycle
                    for well in wells:
                        sheets[steps.index(step)].cell(row=1+cycle,column=2+wells.index(well)).value=float(tempMatrix[well[1]][well[2]].replace(",","."))

    print(nb_cycles)

    #We create the table of average temperatures

    #We put them away fro mthe top and the left
    delta_row=6
    delta_col=2

    for sheet in sheets:
        #We trace numbers of columns
        for col in range(1,12+1):
            sheet.cell(row=nb_cycles+delta_row,column=delta_col+col).value=col
        #We trace letter of rows:
        for row in range(1, 8 + 1):
            sheet.cell(row=nb_cycles+delta_row+row,column=delta_col).value=chr(ord("A")+row-1)

        well_list=range(firstWell,lastWell+1)
        for well in well_list:
            row=nb_cycles+delta_row+((well-1)%8)+1
            col=delta_col + ((well-1)//8)+1
            well_cell=sheet.cell(row=1,column=2+well_list.index(well))
            sheet.cell(row=row,column=col).value="=AVERAGE(" + excelColumn(well_cell.column) + str(2) + ":" + excelColumn(well_cell.column) + str(2+ nb_cycles) + ")"


    #On enlève la première feuille inutilisée
    temp_excel.remove(temp_excel['Sheet'])
    #We save
    temp_excel.save(initial_path+"\\temp_.xlsx")